from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font

wb = Workbook()

def print_label(sheet, cell: str, text, font_size: int, font_family, bold: bool = False):
        font = Font(name=font_family,
                    size=font_size,
                    bold=bold)
        sheet[cell] = text
        sheet[cell].font = font


sheet = wb.active
sheet["A2"].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
sheet["A2"] = "esto es un titulo"
sheet.merge_cells("A2:C2")
sheet["B2"].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
sheet["C2"].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
sheet['A1'] = 7
sheet["A3"].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
sheet["A4"].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
sheet["A8"].border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thick'))




print_label(sheet, 'B4','que ase mi jenteeee lindaaa', 13, 'Arial')

sheet['A5'] = 'informacion nutricional '
sheet['A6'] = 'esto tambien es info sin importancia'

sheet.merge_cells('A5:C5')

# print(type(bloque1))


sheet['A4'] = 10
wb.save("tabla.xlsx")