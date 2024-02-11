from constants import ingredient_codes_list, nutrition_info_model
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment


class Table:
    def __init__(self, product: object, document_route: str, position: str, font_size: int) -> None:
        self.product = product
        self.nutrients = self.product.get_nutritional_data().copy()
        self.std_nutrients = self.product.get_nutritional_std_data().copy()
        self.document_route = document_route
        self.position = position
        self.column_ord = ord(self.position[0])
        self.row_number = int(self.position[1])
        macronutrients = ['Calories','Fat','Poly Unsaturated Fat','Saturated Fat','Trans fat','Carbohydrates','Fiber','Sugar','Added sugar','Protein','Sodium']
        
        self.wb = Workbook()
        self.ws = self.wb.active
        self.font_family = 'Arial'
        self.font_size = font_size
    
    def values_filter(self):
        for nutrient in self.std_nutrients.keys():
            if nutrient == 'Calories' and self.std_nutrients[nutrient] <= 4:
                self.std_nutrients['nutrient'] = 0
                self.nutrients['nutrient'] = 0
            elif nutrient in ['Carbohydrates', 'Sugar', 'Protein', 'Fat', 'Fiber'] and self.std_nutrients[nutrient] <= 0.5:
                self.std_nutrients['nutrient'] = 0
                self.nutrients['nutrient'] = 0
            elif nutrient in ['Cholesterol', 'Sodium'] and self.std_nutrients[nutrient] <= 5:
                self.std_nutrients['nutrient'] = 0
                self.nutrients['nutrient'] = 0
        

    def create_table(self):
        self.print_cells(self.position, self.row_number, self.column_ord)
        self.put_borders(self.position, self.row_number, self.column_ord)
        self.merge_cells(self.position, self.row_number, self.column_ord)        
        self.wb.save("Nutritional_table.xlsx")

    def print_cells(self, position, row_number, column_ord):
        self.print_label(position, "Información nutricional", int(self.font_size*1.3),True, 'center')
        self.print_label(position[0] + str(row_number + 1), f"Tamaño de porción: 1 unidad ({self.product.portion_weight}g)", int(self.font_size))
        self.print_label(position[0] + str(row_number + 2), f"Número de porciones por envase: Aprox. {self.product.portions_number}", int(self.font_size))
        self.print_label(position[0] + str(row_number + 3), "Calorías (kcal)", int(self.font_size)*1.3, True)
        self.print_label(chr(self.column_ord + 1) + str(row_number + 3), "Por 100g", self.font_size, horizontal_alig='center')
        self.print_label(chr(self.column_ord + 2) + str(row_number + 3), "Por porción", self.font_size, horizontal_alig='center')
        #self.print_label(chr(self.column_ord + 1) + str(row_number + 4), f"{int(self.product.get_standar_nutritional_data()['Calories']['amount'])}", int(self.font_size)*1.3, horizontal_alig='center')
        self.print_label(chr(self.column_ord + 2) + str(row_number + 4), f"{int(self.product.get_nutritional_data()['Calories']['amount'])}", self.font_size, horizontal_alig='center')

    def put_borders(self, position, row_number, column_ord):
        self.put_border(self.position, 'thin')
        self.put_border(position[0] + str(row_number + 1), borde_bottom=False)
        self.put_border(position[0] + str(row_number + 2), 'thick', border_top=False)
        self.put_border(position[0] + str(row_number + 3), 'thick')
        self.put_border(chr(self.column_ord + 1) + str(row_number + 3))
        self.put_border(chr(self.column_ord + 2) + str(row_number + 3))
        self.put_border(chr(self.column_ord + 1) + str(row_number + 4), 'thick')
        self.put_border(chr(self.column_ord + 2) + str(row_number + 4), 'thick')

    def merge_cells(self, position, row_number, column_ord):
        self.ws.merge_cells(f"{position}:{chr(column_ord + 2)}{position[1]}")
        self.ws.merge_cells(f"{position[0]}{str(row_number + 1)}:{chr(column_ord + 2)}{str(row_number + 1)}")
        self.ws.merge_cells(f"{position[0]}{str(row_number + 2)}:{chr(column_ord + 2)}{str(row_number + 2)}")
        self.ws.merge_cells(f"{position[0]}{str(row_number + 3)}:{position[0]}{str(row_number + 4)}")

    def print_label(self, cell: str, text, font_size: int, bold: bool = False, horizontal_alig: str = 'left'):
        font = Font(name=self.font_family,
                    size=font_size,
                    bold=bold)
        alignment = Alignment(horizontal=horizontal_alig, 
                              vertical='top')
        self.ws[cell] = text
        self.ws[cell].font = font
        self.ws[cell].alignment = alignment

    def put_border(self, cell, style_bottom: str = 'thin', border_top: bool = True, borde_bottom: bool = True):
        if border_top and borde_bottom:
            bottom_side = Side(style=style_bottom)
            top_side = Side(style='thin')
        elif not border_top and borde_bottom:
            bottom_side = Side(style=style_bottom)
            top_side = None
        elif border_top and not borde_bottom:
            bottom_side = None
            top_side = Side(style='thin')
        else:
            bottom_side = None
            top_side = None
            
        border_cell = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top= top_side, 
                        bottom=bottom_side)

        self.ws[cell].border = border_cell

 


class Table_vertical_standar(Table):
    def __init__(self, product: object, document_route: str, position: str, font_size: int) -> None:
        super().__init__(product, document_route, position, font_size)


    def print_aditional_labels(self):
        pass


class Table_simplified(Table):
    def __init__(self, type: str, product: object) -> None:
        super().__init__(type, product)

    def print_aditional_labels(self):
        pass

    def print_table(self):
        self.print_labels()


class Table_tubular_linear(Table):
    def __init__(self, type: str, product: object) -> None:
        super().__init__(type, product)

    def print_aditional_labels(self):
        pass

    def print_table(self):
        self.print_labels()

    def merge_table_cells(self):
        column_ord = ord(self.position[0])
        row_number = int(self.position[1])
        self.block_title = self.ws.merge_cells(f"{self.position}:{chr(column_ord + 2)}{self.position[1]}")
        self.block_portion_size = self.ws.merge_cells(f"{self.position[0]}{str(row_number + 1)}:{chr(column_ord + 2)}{str(row_number + 1)}")
        self.block_portion_number = self.ws.merge_cells(f"{self.position[0]}{str(row_number + 2)}:{chr(column_ord + 2)}{str(row_number + 2)}")
        self.block_protein = self.ws.merge_cells(f"{self.position[0]}{str(row_number + 3)}:{self.position[0]}{str(row_number + 4)}")



class Product:
    def __init__(self, 
                 name: str,
                 ingredients_input: dict, 
                 portion_weight: float,
                 total_weight: float,
                 nutritional_data: dict = nutrition_info_model,
                 data_from_api: bool = True 
                 ) -> None:
        self.name = name
        self.ingredients_input = ingredients_input
        self.ingredients = []

        self.nutritional_data = nutritional_data
        self.nutritional_std_data = nutrition_info_model

        self.total_weight = total_weight
        self.portion_weight = portion_weight
        self.portions_number = int(self.total_weight / self.portion_weight)

        self.data_from_api = data_from_api

    def create_ingredients_list(self, api: object) -> list: 
        for ingredient in self.ingredients_input:
            ingredient_obj =Ingredient(ingredient[0], ingredient[1]*self.portion_weight)
            ingredient_obj.search_nutritional_data(api)
            self.ingredients.append(ingredient_obj)

    def generate_nutritional_data(self) -> None:
        if self.data_from_api:
            for index in range(len(self.ingredients)):
                for nutrient in self.nutritional_data.keys():
                    if nutrient in self.ingredients[index].get_nutritional_data().keys():
                        self.nutritional_data[nutrient]['amount'] = self.nutritional_data[nutrient]['amount'] +  self.ingredients[index].get_nutritional_data()[nutrient]['amount']
                        self.nutritional_data[nutrient]['percentOfDailyNeeds'] = self.nutritional_data[nutrient]['percentOfDailyNeeds'] + self.ingredients[index].get_nutritional_data()[nutrient]['percentOfDailyNeeds']
                    else:
                        continue
        else:
            pass
    
    def generate_nutritional_std_data(self):
        for nutrient in self.nutritional_data.keys():
            self.nutritional_std_data[nutrient]['amount'] = round(100*(self.nutritional_data[nutrient]['amount'] / self.portion_weight),2)
            self.nutritional_std_data[nutrient]['percentOfDailyNeeds'] = round(100*(self.nutritional_data[nutrient]['percentOfDailyNeeds'] / self.portion_weight),2)

    def generate_all_data(self):
        self.generate_nutritional_data()
        self.generate_nutritional_std_data()

    def added_sugar(self, name: str, amount_percent: float):
        sugar_portion = self.portion_weight*amount_percent
        sugar_standar = 100*amount_percent
        self.ingredients.append(Ingredient(name, sugar_portion))
        self.ingredients.append(Ingredient(name, sugar_standar))

        self.nutritional_data['Added sugar']['amount'] = sugar_portion
        self.nutritional_data['Calories']['amount'] = self.nutritional_data['Calories']['amount'] + sugar_portion*4
        self.nutritional_std_data['Added sugar']['amount'] = sugar_standar 
        self.nutritional_std_data['Calories']['amount'] = self.nutritional_std_data['Calories']['amount'] + sugar_standar*4

    def get_nutritional_data(self) -> dict:
        return self.nutritional_data
    
    def get_nutritional_std_data(self) -> dict:
        return self.nutritional_std_data
    
    def get_portions_number(self):
        return self.portions_number


class Ingredient:
    def __init__(self, name: str, amount: float) -> None:
        self.name = name
        self.amount = amount #the amount of ingredient in grams 
        self.nutritional_data = {}
    
    def search_nutritional_data(self, api: object) -> None:
        id = ingredient_codes_list[self.name] # search the code of the ingredien in ingredient_code_list dictionary into constans.py
        api_nutritional_data = api.do_requests(id, added_params={'amount': str(self.amount), 'unit': 'grams'})['nutrition']['nutrients']
        # We use the api object to do the consult
        for nutrient in api_nutritional_data:
            self.nutritional_data[nutrient['name']] = {'amount': nutrient['amount'], 
                                                       'unit': nutrient['unit'], 
                                                       'percentOfDailyNeeds': nutrient['percentOfDailyNeeds']}

    def get_nutritional_data(self) -> dict:
        return self.nutritional_data

    def get_nutrient(self, nutrient: str) -> float:
        return self.nutritional_data[nutrient]